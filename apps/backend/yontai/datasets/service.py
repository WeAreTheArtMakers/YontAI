from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any
from urllib.parse import quote

import httpx
from sqlalchemy.orm import Session

from yontai.core.paths import storage_path
from yontai.db.models import Dataset
from yontai.repositories.datasets import DatasetRepository
from yontai.schemas.datasets import DatasetCreate, PublicDatasetCatalogItem, PublicDatasetImport

PUBLIC_DATASET_CATALOG = [
    PublicDatasetCatalogItem(
        repository_id="TFLai/Turkish-Alpaca",
        title="Turkish Alpaca",
        task_type="instruction",
        language="tr",
        license="apache-2.0",
        description_tr=(
            "Türkçe instruction-response fine-tuning denemeleri için "
            "Alpaca biçimli veri seti."
        ),
        recommended_limit=1000,
    ),
    PublicDatasetCatalogItem(
        repository_id="cgulse/alpaca-cleaned-tr",
        title="Alpaca Cleaned TR",
        task_type="instruction",
        language="tr",
        license="cc-by-4.0",
        description_tr=(
            "JSON formatlı, yaklaşık 40 bin satırlık temizlenmiş Türkçe Alpaca "
            "instruction veri seti. Hızlı örnekleme ve küçük fine-tuning denemeleri için uygun."
        ),
        recommended_limit=1000,
    ),
    PublicDatasetCatalogItem(
        repository_id="BrewInteractive/alpaca-tr",
        title="BrewInteractive Alpaca TR",
        task_type="instruction",
        language="tr",
        license="apache-2.0",
        description_tr=(
            "Yaklaşık 45 bin satırlık Türkçe Alpaca türevi instruction veri seti. "
            "Parquet formatlıdır; YontAI örneklemeyi satır limitiyle yapar."
        ),
        recommended_limit=1000,
    ),
    PublicDatasetCatalogItem(
        repository_id="tascib/turkish-instruction",
        title="Turkish Instruction",
        task_type="instruction",
        language="tr",
        license=None,
        description_tr=(
            "Birden fazla açık kaynaktan temizlenmiş Türkçe instruction-tuning derlemesi."
        ),
        recommended_limit=1000,
    ),
    PublicDatasetCatalogItem(
        repository_id="hcsolakoglu/turkish-wikipedia-qa-4-million",
        title="Turkish Wikipedia QA",
        task_type="qa",
        language="tr",
        license="cc-by-nc-sa-4.0",
        description_tr=(
            "Türkçe Wikipedia kaynaklı geniş ölçekli soru-cevap veri seti. "
            "Araştırma ve ticari olmayan kullanım için uygundur; dosya boyutu büyüktür."
        ),
        recommended_limit=500,
    ),
    PublicDatasetCatalogItem(
        repository_id="merve/turkish_instructions",
        title="Turkish Instructions",
        task_type="instruction",
        language="tr",
        license="apache-2.0",
        description_tr="Türkçe komut-yanıt görevleri için yaygın kullanılan açık veri seti.",
        recommended_limit=1000,
    ),
]


class DatasetRegistryService:
    def __init__(self, db: Session) -> None:
        self.repo = DatasetRepository(db)

    def list_datasets(self) -> list[Dataset]:
        return self.repo.list()

    def get_dataset(self, dataset_id: str) -> Dataset | None:
        return self.repo.get(dataset_id)

    def register_from_path(self, payload: DatasetCreate) -> Dataset:
        path = Path(payload.path).expanduser().resolve()
        if not path.exists():
            raise ValueError(f"Veri seti dosyası bulunamadı: {path}")
        rows = self._read_dataset(path, payload.format)
        analysis = self._analyze_rows(rows)
        dataset = Dataset(
            project_id=payload.project_id,
            name=payload.name.strip(),
            source_type="local_file",
            path=str(path),
            format=payload.format,
            task_type=payload.task_type or self._infer_task_type(rows),
            **analysis,
        )
        return self.repo.add(dataset)

    def register_upload(
        self,
        *,
        filename: str,
        content: bytes,
        name: str | None = None,
        project_id: str | None = None,
        task_type: str | None = None,
    ) -> Dataset:
        dataset_format = self._format_from_filename(filename)
        uploads_dir = storage_path("datasets") / "uploads"
        uploads_dir.mkdir(parents=True, exist_ok=True)
        digest = hashlib.sha256(content).hexdigest()[:16]
        safe_name = Path(filename).name.replace(" ", "_")
        target = uploads_dir / f"{digest}_{safe_name}"
        target.write_bytes(content)
        return self.register_from_path(
            DatasetCreate(
                name=name or Path(filename).stem,
                path=str(target),
                format=dataset_format,
                project_id=project_id,
                task_type=task_type,
            )
        )

    def public_catalog(self) -> list[PublicDatasetCatalogItem]:
        return PUBLIC_DATASET_CATALOG

    async def import_public_dataset(self, payload: PublicDatasetImport) -> Dataset:
        repository_id = payload.repository_id.strip()
        file_name, dataset_format = await self._select_huggingface_dataset_file(repository_id)
        rows = await self._download_huggingface_sample(
            repository_id=repository_id,
            file_name=file_name,
            dataset_format=dataset_format,
            max_rows=payload.max_rows,
        )
        if not rows:
            raise ValueError("Public veri setinden örnek kayıt okunamadı.")

        public_dir = storage_path("datasets") / "public"
        public_dir.mkdir(parents=True, exist_ok=True)
        safe_repo = repository_id.replace("/", "__").replace(" ", "_")
        target = self._unique_target(
            public_dir / f"{safe_repo}_{payload.max_rows}.{dataset_format}"
        )
        self._write_dataset(target, dataset_format, rows)

        dataset = self.register_from_path(
            DatasetCreate(
                name=payload.name or f"{repository_id} ({len(rows)} örnek)",
                path=str(target),
                format=dataset_format,
                project_id=payload.project_id,
                task_type=(
                    payload.task_type
                    or self._catalog_task_type(repository_id)
                    or "instruction"
                ),
            )
        )
        dataset.source_type = "huggingface_public"
        dataset.statistics = {
            **(dataset.statistics or {}),
            "public_source": {
                "repository_id": repository_id,
                "file": file_name,
                "sampled_rows": len(rows),
                "max_rows": payload.max_rows,
                "source_url": (
                    f"https://huggingface.co/datasets/{repository_id}/resolve/main/{file_name}"
                ),
            },
        }
        return self.repo.save(dataset)

    def analyze(self, dataset_id: str) -> Dataset | None:
        dataset = self.repo.get(dataset_id)
        if dataset is None:
            return None
        rows = self._read_dataset(Path(dataset.path), dataset.format)
        analysis = self._analyze_rows(rows)
        for key, value in analysis.items():
            setattr(dataset, key, value)
        dataset.task_type = dataset.task_type or self._infer_task_type(rows)
        return self.repo.save(dataset)

    def create_cleaned_dataset(self, dataset_id: str, action: str) -> Dataset | None:
        dataset = self.repo.get(dataset_id)
        if dataset is None:
            return None

        rows = self._read_dataset(Path(dataset.path), dataset.format)
        if action == "remove_duplicates":
            cleaned_rows = self._remove_duplicate_rows(rows)
            suffix = "tekrarlar_temizlendi"
            title = "Tekrarlar Temizlendi"
        elif action == "remove_low_quality":
            cleaned_rows = [
                row
                for row in rows
                if not self._is_empty(row) and self._estimate_tokens(row) >= 8
            ]
            suffix = "dusuk_kalite_temizlendi"
            title = "Düşük Kalite Temizlendi"
        else:
            raise ValueError(f"Desteklenmeyen veri düzeltme aksiyonu: {action}")

        if len(cleaned_rows) == len(rows):
            return dataset

        cleaned_dir = storage_path("datasets") / "cleaned"
        cleaned_dir.mkdir(parents=True, exist_ok=True)
        source_path = Path(dataset.path)
        target = self._unique_target(
            cleaned_dir / f"{source_path.stem}_{suffix}{source_path.suffix}"
        )
        self._write_dataset(target, dataset.format, cleaned_rows)
        return self.register_from_path(
            DatasetCreate(
                name=f"{dataset.name} - {title}",
                path=str(target),
                format=dataset.format,
                project_id=dataset.project_id,
                task_type=dataset.task_type,
            )
        )

    def _catalog_task_type(self, repository_id: str) -> str | None:
        for item in PUBLIC_DATASET_CATALOG:
            if item.repository_id == repository_id:
                return item.task_type
        return None

    async def _select_huggingface_dataset_file(self, repository_id: str) -> tuple[str, str]:
        encoded_repo = quote(repository_id, safe="/")
        url = f"https://huggingface.co/api/datasets/{encoded_repo}"
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            response = await client.get(url)
            response.raise_for_status()
            payload = response.json()

        supported = {
            ".jsonl": "jsonl",
            ".json": "json",
            ".parquet": "parquet",
            ".csv": "csv",
            ".txt": "txt",
        }
        siblings = payload.get("siblings") or []
        candidates: list[tuple[str, str]] = []
        for sibling in siblings:
            file_name = str(sibling.get("rfilename") or "")
            lowered = file_name.lower()
            if file_name.startswith(".") or "/.git" in file_name:
                continue
            for suffix, dataset_format in supported.items():
                if lowered.endswith(suffix):
                    candidates.append((file_name, dataset_format))
                    break

        if not candidates:
            raise ValueError(
                "Bu Hugging Face veri setinde JSON, JSONL, Parquet, CSV veya "
                "TXT dosyası bulunamadı."
            )

        def priority(candidate: tuple[str, str]) -> tuple[int, str]:
            file_name, dataset_format = candidate
            format_rank = {
                "jsonl": 0,
                "json": 1,
                "parquet": 2,
                "csv": 3,
                "txt": 4,
            }[dataset_format]
            split_rank = 0 if any(part in file_name.lower() for part in ("train", "data")) else 1
            return (format_rank + split_rank, file_name)

        return sorted(candidates, key=priority)[0]

    async def _download_huggingface_sample(
        self,
        *,
        repository_id: str,
        file_name: str,
        dataset_format: str,
        max_rows: int,
    ) -> list[dict[str, Any]]:
        encoded_repo = quote(repository_id, safe="/")
        encoded_file = quote(file_name, safe="/")
        url = f"https://huggingface.co/datasets/{encoded_repo}/resolve/main/{encoded_file}"
        async with httpx.AsyncClient(timeout=90.0, follow_redirects=True) as client:
            response = await client.get(url)
            response.raise_for_status()
            content = response.content
        if len(content) > 80 * 1024 * 1024:
            raise ValueError(
                "Veri dosyası hızlı örnekleme için çok büyük; "
                "daha küçük bir dosya seçin."
            )

        sample_dir = storage_path("datasets") / "tmp"
        sample_dir.mkdir(parents=True, exist_ok=True)
        sample_path = sample_dir / f"public_sample.{dataset_format}"

        if dataset_format == "json":
            rows = self._rows_from_json_content(content)[:max_rows]
            return [self._normalize_record(row) for row in rows]

        if dataset_format == "parquet":
            sample_path.write_bytes(content)
            return self._read_parquet(sample_path, max_rows=max_rows)

        text = content.decode("utf-8-sig", errors="replace")
        if dataset_format == "jsonl":
            lines = [line for line in text.splitlines() if line.strip()][:max_rows]
            sample_path.write_text("\n".join(lines), encoding="utf-8")
            return self._read_jsonl(sample_path)
        if dataset_format == "csv":
            lines = text.splitlines()
            if len(lines) > max_rows + 1:
                text = "\n".join(lines[: max_rows + 1])
            sample_path.write_text(text, encoding="utf-8")
            return self._read_csv(sample_path)
        if dataset_format == "txt":
            sample_path.write_text("\n".join(text.splitlines()[:max_rows]), encoding="utf-8")
            return self._read_txt(sample_path)
        raise ValueError(f"Desteklenmeyen public veri formatı: {dataset_format}")

    def _rows_from_json_content(self, content: bytes) -> list[Any]:
        payload = json.loads(content.decode("utf-8-sig"))
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            for value in payload.values():
                if isinstance(value, list):
                    return value
            return [payload]
        raise ValueError("Public JSON veri seti liste veya nesne içermelidir.")

    def _format_from_filename(self, filename: str) -> str:
        suffix = Path(filename).suffix.lower().lstrip(".")
        if suffix not in {"json", "jsonl", "csv", "xlsx", "txt", "parquet"}:
            raise ValueError(
                "Desteklenen veri seti formatları: JSON, JSONL, CSV, XLSX, TXT, Parquet."
            )
        return suffix

    def _read_dataset(self, path: Path, dataset_format: str) -> list[dict[str, Any]]:
        if dataset_format == "json":
            return self._read_json(path)
        if dataset_format == "jsonl":
            return self._read_jsonl(path)
        if dataset_format == "csv":
            return self._read_csv(path)
        if dataset_format == "xlsx":
            return self._read_xlsx(path)
        if dataset_format == "txt":
            return self._read_txt(path)
        if dataset_format == "parquet":
            return self._read_parquet(path)
        raise ValueError(f"Desteklenmeyen format: {dataset_format}")

    def _read_json(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8") as file:
            payload = json.load(file)
        if isinstance(payload, list):
            return [self._normalize_record(item) for item in payload]
        if isinstance(payload, dict):
            for value in payload.values():
                if isinstance(value, list):
                    return [self._normalize_record(item) for item in value]
            return [self._normalize_record(payload)]
        raise ValueError("JSON veri seti liste veya nesne içermelidir.")

    def _read_jsonl(self, path: Path) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        with path.open("r", encoding="utf-8") as file:
            for line_no, line in enumerate(file, start=1):
                stripped = line.strip()
                if not stripped:
                    rows.append({"text": ""})
                    continue
                try:
                    rows.append(self._normalize_record(json.loads(stripped)))
                except json.JSONDecodeError as exc:
                    raise ValueError(f"JSONL satır {line_no} geçerli JSON değil.") from exc
        return rows

    def _read_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            return [dict(row) for row in csv.DictReader(file)]

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("XLSX desteği için openpyxl bağımlılığı gereklidir.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [
            str(value) if value is not None else f"kolon_{index + 1}"
            for index, value in enumerate(rows[0])
        ]
        records = []
        for row in rows[1:]:
            records.append(
                {
                    headers[index]: row[index] if index < len(row) else None
                    for index in range(len(headers))
                }
            )
        return records

    def _read_txt(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8") as file:
            return [{"text": line.rstrip("\n")} for line in file]

    def _read_parquet(self, path: Path, max_rows: int | None = None) -> list[dict[str, Any]]:
        try:
            import pandas as pd
        except ImportError as exc:
            raise ValueError("Parquet okuma desteği için pandas/pyarrow gereklidir.") from exc

        dataframe = pd.read_parquet(path)
        if max_rows is not None:
            dataframe = dataframe.head(max_rows)
        dataframe = dataframe.where(pd.notnull(dataframe), None)
        return [self._normalize_record(row) for row in dataframe.to_dict(orient="records")]

    def _write_dataset(
        self,
        path: Path,
        dataset_format: str,
        rows: list[dict[str, Any]],
    ) -> None:
        if dataset_format == "json":
            path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
            return
        if dataset_format == "jsonl":
            with path.open("w", encoding="utf-8") as file:
                for row in rows:
                    file.write(json.dumps(row, ensure_ascii=False) + "\n")
            return
        if dataset_format == "csv":
            fieldnames = sorted({key for row in rows for key in row})
            with path.open("w", encoding="utf-8", newline="") as file:
                writer = csv.DictWriter(file, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
            return
        if dataset_format == "txt":
            with path.open("w", encoding="utf-8") as file:
                for row in rows:
                    file.write(str(row.get("text", "")) + "\n")
            return
        if dataset_format == "xlsx":
            self._write_xlsx(path, rows)
            return
        if dataset_format == "parquet":
            self._write_parquet(path, rows)
            return
        raise ValueError(f"Desteklenmeyen format: {dataset_format}")

    def _write_parquet(self, path: Path, rows: list[dict[str, Any]]) -> None:
        try:
            import pandas as pd
        except ImportError as exc:
            raise ValueError("Parquet yazma desteği için pandas/pyarrow gereklidir.") from exc
        pd.DataFrame(rows).to_parquet(path, index=False)

    def _write_xlsx(self, path: Path, rows: list[dict[str, Any]]) -> None:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise ValueError("XLSX yazma desteği için openpyxl bağımlılığı gereklidir.") from exc

        workbook = Workbook()
        sheet = workbook.active
        headers = sorted({key for row in rows for key in row})
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        workbook.save(path)

    def _remove_duplicate_rows(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        seen: set[str] = set()
        cleaned: list[dict[str, Any]] = []
        for row in rows:
            normalized = json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
            if normalized in seen:
                continue
            seen.add(normalized)
            cleaned.append(row)
        return cleaned

    def _unique_target(self, target: Path) -> Path:
        if not target.exists():
            return target
        for index in range(1, 10_000):
            candidate = target.with_name(f"{target.stem}_{index}{target.suffix}")
            if not candidate.exists():
                return candidate
        raise ValueError("Benzersiz veri seti dosya adı üretilemedi.")

    def _normalize_record(self, item: Any) -> dict[str, Any]:
        if isinstance(item, dict):
            return item
        return {"text": item}

    def _analyze_rows(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        row_count = len(rows)
        normalized = [
            json.dumps(row, ensure_ascii=False, sort_keys=True, default=str) for row in rows
        ]
        duplicate_count = sum(count - 1 for count in Counter(normalized).values() if count > 1)
        empty_count = sum(1 for row in rows if self._is_empty(row))
        token_counts = [self._estimate_tokens(row) for row in rows]
        token_total = sum(token_counts)
        duplicate_ratio = duplicate_count / row_count if row_count else 0
        empty_ratio = empty_count / row_count if row_count else 0
        average_tokens = token_total / row_count if row_count else 0
        schema = self._infer_schema(rows)
        quality_score = self._quality_score(row_count, duplicate_ratio, empty_ratio, average_tokens)
        report = self._build_report(
            row_count, duplicate_ratio, empty_ratio, average_tokens, quality_score
        )
        # Dataset Intelligence Features
        tr_count = 0
        en_count = 0
        short_examples = 0
        long_examples = 0
        quality_distribution = {"Yüksek": 0, "Orta": 0, "Düşük": 0}
        histogram_bins = {"0-64": 0, "65-256": 0, "257-1024": 0, "1024+": 0}
        
        for i, row in enumerate(rows):
            text = " ".join("" if value is None else str(value) for value in row.values()).lower()
            if any(w in text for w in [" ve ", " bir ", " için ", " bu ", " çok "]):
                tr_count += 1
            elif any(w in text for w in [" the ", " and ", " is ", " to ", " a "]):
                en_count += 1

            tokens = token_counts[i]
            if tokens < 8:
                short_examples += 1
            elif tokens > 1024:
                long_examples += 1

            if tokens <= 64:
                histogram_bins["0-64"] += 1
            elif tokens <= 256:
                histogram_bins["65-256"] += 1
            elif tokens <= 1024:
                histogram_bins["257-1024"] += 1
            else:
                histogram_bins["1024+"] += 1

            row_score = 100
            if self._is_empty(row):
                row_score -= 50
            if tokens < 8:
                row_score -= 30
            if row_score >= 80:
                quality_distribution["Yüksek"] += 1
            elif row_score >= 50:
                quality_distribution["Orta"] += 1
            else:
                quality_distribution["Düşük"] += 1

        for k in quality_distribution:
            quality_distribution[k] = (
                round((quality_distribution[k] / row_count * 100), 1) if row_count else 0
            )

        tr_percentage = round((tr_count / row_count * 100), 1) if row_count else 0
        en_percentage = round((en_count / row_count * 100), 1) if row_count else 0
        other_percentage = round(100 - tr_percentage - en_percentage, 1) if row_count else 0

        return {
            "row_count": row_count,
            "token_count_estimate": token_total,
            "average_tokens": round(average_tokens, 2),
            "duplicate_ratio": round(duplicate_ratio, 4),
            "empty_ratio": round(empty_ratio, 4),
            "quality_score": quality_score,
            "dataset_schema": schema,
            "preview": rows[:20],
            "statistics": {
                "duplicate_rows": duplicate_count,
                "empty_records": empty_count,
                "min_tokens": min(token_counts) if token_counts else 0,
                "max_tokens": max(token_counts) if token_counts else 0,
                "avg_tokens": round(average_tokens, 2),
                "dominant_language": "tr" if tr_count >= en_count else "en",
                "language_distribution": {
                    "Türkçe": tr_percentage,
                    "İngilizce": en_percentage,
                    "Diğer": max(0, other_percentage),
                },
                "quality_distribution": quality_distribution,
                "token_histogram": histogram_bins,
                "short_examples_count": short_examples,
                "long_examples_count": long_examples,
            },
            "report": report,
        }

    def _is_empty(self, row: dict[str, Any]) -> bool:
        return all(value is None or str(value).strip() == "" for value in row.values())

    def _estimate_tokens(self, row: dict[str, Any]) -> int:
        text = " ".join("" if value is None else str(value) for value in row.values())
        pieces = re.findall(r"\w+|[^\w\s]", text, flags=re.UNICODE)
        return max(1, int(len(pieces) * 1.15)) if text.strip() else 0

    def _infer_schema(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        columns: dict[str, Counter[str]] = {}
        for row in rows[:1000]:
            for key, value in row.items():
                columns.setdefault(key, Counter())[self._type_name(value)] += 1
        return {
            "columns": [
                {"name": key, "types": dict(counter), "primary_type": counter.most_common(1)[0][0]}
                for key, counter in columns.items()
            ]
        }

    def _type_name(self, value: Any) -> str:
        if value is None or value == "":
            return "empty"
        if isinstance(value, bool):
            return "bool"
        if isinstance(value, int):
            return "int"
        if isinstance(value, float):
            return "float"
        return "text"

    def _infer_task_type(self, rows: list[dict[str, Any]]) -> str:
        keys = {key.lower() for row in rows[:50] for key in row.keys()}
        if {"messages"} & keys:
            return "chat"
        if {"prompt", "completion"} <= keys or {"instruction", "output"} <= keys:
            return "instruction"
        if {"label", "text"} <= keys:
            return "classification"
        return "raw_text"

    def _quality_score(
        self, row_count: int, duplicate_ratio: float, empty_ratio: float, average_tokens: float
    ) -> float:
        score = 100.0
        score -= duplicate_ratio * 35
        score -= empty_ratio * 35
        if row_count < 500:
            score -= 25
        elif row_count < 5_000:
            score -= 10
        if average_tokens < 8:
            score -= 10
        return round(max(0, min(100, score)), 1)

    def _build_report(
        self,
        row_count: int,
        duplicate_ratio: float,
        empty_ratio: float,
        average_tokens: float,
        score: float,
    ) -> dict[str, Any]:
        findings: list[str] = []
        if row_count < 500:
            findings.append("Örnek sayısı düşük; fine-tuning sonucu kararsız olabilir.")
        if duplicate_ratio > 0.1:
            findings.append("Tekrar eden kayıt oranı yüksek; ezberleme riski artar.")
        if empty_ratio > 0.05:
            findings.append("Boş kayıt oranı temizleme gerektiriyor.")
        if average_tokens < 8:
            findings.append("Ortalama örnek uzunluğu düşük; görev sinyali zayıf olabilir.")
        return {
            "summary_tr": (
                f"Veri seti kalite skoru {score}/100. Toplam {row_count} örnek analiz edildi."
            ),
            "findings": findings or ["Belirgin veri kalitesi sorunu tespit edilmedi."],
            "recommended_actions": self._recommended_actions(
                row_count, duplicate_ratio, empty_ratio
            ),
        }

    def _recommended_actions(
        self, row_count: int, duplicate_ratio: float, empty_ratio: float
    ) -> list[str]:
        actions: list[str] = []
        if row_count < 5_000:
            actions.append("Fine-tuning için en az 5000 kaliteli örnek hedefleyin.")
        if duplicate_ratio > 0:
            actions.append("Aynı içeriğe sahip kayıtları eğitimden önce temizleyin.")
        if empty_ratio > 0:
            actions.append("Boş veya eksik kayıtları doğrulama aşamasında ayırın.")
        return actions or ["Mevcut veri setiyle benchmark veya küçük ölçekli deneme yapılabilir."]
